import os
import random
from difflib import SequenceMatcher

import pandas as pd

from openpyxl import styles, load_workbook, formatting
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle


def cleanup_results_dir(results_dir):
    for directory in os.listdir(results_dir):
        directory = f"{results_dir}/{directory}"
        for file in os.listdir(directory):
            if file.split('.')[1] == 'csv':
                file_path = f"{directory}/{file}"
                df = pd.read_csv(file_path)
                if file not in ['control_smiles.csv', 'hyer_smiles.csv', 'jaccard_smiles.csv', 'rdkit_smiles.csv']:
                    df = df.sort_values(by=['smiles'])
                else:
                    df = df.sort_values(by=['smiles', 'pred_error'])
                df.to_csv(file_path, index=False)


def gem_sequence_ratio(results_dir):
    for directory in os.listdir(results_dir):
        directory = f"{results_dir}/{directory}"
        ratio_df = []

        control_df = pd.read_csv(f"{directory}/control_smiles.csv")
        control_models_order = control_df['run_name'].tolist()

        compare_df = pd.read_csv(f"{directory}/compare.csv")

        files = ['rdkit', 'hyer', 'jaccard']
        for file in files:
            file_df = pd.read_csv(f"{directory}/{file}_smiles.csv")
            file_smiles = list(set(file_df['smiles'].tolist()))
            for file_smile in file_smiles:
                file_smile_model_order = file_df.loc[file_df['smiles'] == file_smile]
                file_smile_model_order = file_smile_model_order['run_name'].tolist()
                sr = SequenceMatcher(None, control_models_order, file_smile_model_order).ratio()
                sim_scores = compare_df.loc[compare_df['smiles'] == file_smile].to_dict('records')[0]
                ratio_df.append({'type': file, 'smiles': file_smile, 'ratio_score': sr,
                                 'rdkit_similarity': sim_scores['rdkit_similarity'],
                                 'hyer_similarity': sim_scores['hyer_similarity'],
                                 'jaccard_similarity': sim_scores['jaccard_similarity'],
                                 'percent_error_rdkit_hyer': sim_scores['percent_error_rdkit_hyer'],
                                 'percent_error_rdkit_jaccard': sim_scores['percent_error_rdkit_jaccard']})

        ratio_df = pd.DataFrame(ratio_df)
        ratio_df.to_csv(f"{directory}/ratio.csv")


def apply_conditional_formating(results_dir):
    for directory in os.listdir(results_dir):
        raw_directory = directory
        directory = f"{results_dir}/{directory}"
        writer = pd.ExcelWriter(f"{directory}/{raw_directory}_results.xlsx")
        for file in os.listdir(directory):
            if file.split('.')[1] == 'csv':
                file_path = f"{directory}/{file}"
                file_name = file.split('.')[0]
                file_df = pd.read_csv(file_path)
                file_df.to_excel(writer, sheet_name=file_name, index=False)
        writer.save()

        COLOR_INDEX = [
            '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',  # 0-4
            '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',  # 5-9
            '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
            '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
            '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
            '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
            '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
            '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
            '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
            '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
            '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
            '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
            '00993300', '00993366', '00333399', '00333333',  # 60-63
        ]

        control_models = pd.read_csv(f"{directory}/control_smiles.csv")
        control_models = control_models['run_name'].tolist()
        control_color_dict = {}
        for control_model in control_models:
            hex_number = random.choice(COLOR_INDEX)
            while hex_number in control_color_dict.values():
                hex_number = random.choice(COLOR_INDEX)
            control_color_dict[control_model] = hex_number

        worksheets = ['rdkit_smiles', 'hyer_smiles', 'jaccard_smiles']
        wb = load_workbook(f"{directory}/{raw_directory}_results.xlsx")
        for model, hex_number in control_color_dict.items():
            for worksheet in worksheets:
                worksheet = wb[worksheet]
                red_fill = PatternFill(bgColor=hex_number)
                dxf = DifferentialStyle(fill=red_fill)
                rule = Rule(type="containsText", operator="containsText", text=model, dxf=dxf)
                rule.formula = [f'NOT(ISERROR(SEARCH("{model}",A1)))']
                worksheet.conditional_formatting.add('A1:F1000', rule)

        wb.save(f"{directory}/{raw_directory}_results.xlsx")


if __name__ == "__main__":
    # cleanup_results_dir('results')
    # gem_sequence_ratio('results')
    apply_conditional_formating('results')
